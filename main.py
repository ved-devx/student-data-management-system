from tkinter import *
from tkinter import ttk
from tkinter import messagebox as msg
from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import filedialog as flg
import sys

global file_save_state
file_save_state=False
save_file_dialog = None
wb=None

def add_student():
    global file_save_state
    file_save_state = False 
    name_val = name.get().strip()
    age_raw = age.get().strip()
    Class_raw = Class.get().strip()
    
    if not name_val:
        msg.showerror("Empty Field" , "Please enter name first !")
        return
    
    if not all(x.isalpha() or x.isspace() for x in name_val):
        msg.showerror("Value Error" , "Name should not contain numeric values")
        return

    try:
        if not age_raw or not Class_raw:
            msg.showerror("Empty Field" , "Please enter age and class first !")
            return

        age_ = int(age.get())
        Class_ = int(Class.get())

        if age_ <= 0 or Class_ <= 0:
            msg.showerror("Wrong Entry" , "Please enter the correct age!(Greater than 0)")
            return
    except ValueError:
        msg.showerror("Invalid Age and Class" , "Age and Class must be Numeric Values")
        return
    

    age_Value = int(age.get()) 
    Class_value = int(Class.get())

    if Class_value == 1:
        Class_value = f"{Class_value}st"
    elif Class_value == 2:
        Class_value = f"{Class_value}nd"
    elif Class_value == 3:
        Class_value = f"{Class_value}rd"
    else:
        Class_value = f"{Class_value}th"

    data = (name.get().strip().upper() , age_Value , Class_value)

    duplicate_found = False
    for item in t1.get_children():
        existing = t1.item(item)["values"]
        
        # Convert both to lists for comparison
        if [str(x)for x in existing] == [str(x) for x in data]:
            duplicate_found = True
            break
        
    if duplicate_found:
        duplicate_data = msg.askyesno("Duplicate Data", "Do you want to enter a duplicate entry?")
        if not duplicate_data:
            return

    t1.insert("", END, values=data)
    
    name.set("")
    age.set("")
    Class.set("")


def delete_student():
    global file_save_state
    file_save_state = False
    for i in t1.selection():
        t1.delete(i)

def export_to_excel():
    # Create an workbook
    global wb
    global save_file_dialog
    wb = Workbook()
    ws = wb.active
    ws.title = "Students Data"

    ws.append(["Name" , "Age" , "Class"])

    for items in t1.get_children():
        values = t1.item(items)["values"]
        ws.append(values)

    save_file_dialog =flg.asksaveasfilename(defaultextension=".xlsx" , filetypes=[("Excel Files" , "*.xlsx") , ("All files" , "*.*")])

    if save_file_dialog:
        wb.save(save_file_dialog)
        msg.showinfo("Data Saved" , f"Data Exported to excel on path {save_file_dialog}.")
    

def save():
    global save_file_dialog , wb
    global file_save_state
    if not save_file_dialog:
        export_to_excel()
        if not save_file_dialog:
            return
    
    if save_file_dialog:
        ws = wb.active
        ws.delete_rows(2 , ws.max_row)

        for items in t1.get_children():
            values = t1.item(items)["values"]
            ws.append(values)
        try:
            wb.save(save_file_dialog)
            msg.showinfo("Saved" , f"Data is successfully saved to {save_file_dialog}!")
            file_save_state = True
        except PermissionError:
            msg.showerror("Permission Error", "Please close the excel file before saving!")
    else:
        msg.showwarning("Save" , "Please export to excel first!")

def exit_app():
    global file_save_state
    if file_save_state == True:
        root.destroy()
    else:
        exit_without_save = msg.askyesnocancel("Save" , "Do you want to exit without saving ?")
        if exit_without_save == True:
            root.destroy()
        elif exit_without_save == False:
            save()
            if file_save_state:
                root.destroy()
        else:
            pass


def clear_all():
    clear_all = msg.askyesno("Clear All" , "Are you sure you want to clear full Table?")
    if clear_all == True:
        for items in t1.get_children():
            t1.delete(items)
    else:
        pass


root = Tk()
root.geometry("615x600")
root.maxsize(615 , 600)
root.minsize(615 ,600)
root.title("Students Data")

# Heading
l_head = Label(root , text="Students Data" , font=("comic San MS" , 15 , "bold"))

# Treeview or table
t1 = ttk.Treeview(root , columns=("name" , "age" , "class") , show="headings" , height=15)

t1.heading("name" , text="Name")
t1.heading("age" , text="Age")
t1.heading("class" , text="Class")

# Taking data from user (Entries , labels , tkinter variables) / Details LabelFrame
name = StringVar()
age = StringVar()
Class = StringVar()


lf_details = LabelFrame(root , text="Details")
l1 = Label(lf_details , text="Name " , font=("Roboto" , 15, 'italic'))
l2 = Label(lf_details , text="Age " , font=("Roboto" , 15, 'italic'))
l3 = Label(lf_details , text="Class " , font=("Roboto" , 15, 'italic'))

e_name = Entry(lf_details , textvariable=name)
e_age = Entry(lf_details , textvariable=age)
e_Class = Entry(lf_details , textvariable=Class)

# Functions LabelFrame
lf_functions = LabelFrame(root , text="Edit Functions")
btn1 = Button(lf_functions , text="Add Student" , command=add_student)
btn2 = Button(lf_functions , text="Delete Student" , command=delete_student)
btn3 = Button(lf_functions , text="Export To excel" , command=export_to_excel)

lf_edit_func = LabelFrame(root , text="App Functions")
btn4 = Button(lf_edit_func , text="Save" , command=save)
btn5 = Button(lf_edit_func , text="Exit App" , command=exit_app)
btn6 = Button(lf_edit_func , text="Clear All" ,command=clear_all)

# Placing everything using grid method.
l_head.grid(column=0 , row=0)
t1.grid(column=0 , row=1 , padx=5 , pady=5)
lf_details.grid(row=2 , column=0 , sticky="NW" , padx=5)
lf_functions.grid(row=2 , column=0 , sticky="N")
lf_edit_func.grid(row=2 , column=0 , sticky="NE" , padx=150)

l1.grid(row=0 , column=0 , pady=10)
l2.grid(row=1 , column=0 , pady=10)
l3.grid(row=2 , column=0 , pady=10)

e_name.grid(row=0 , column=1 , padx=5)
e_age.grid(row=1 , column=1 , padx=5)
e_Class.grid(row=2 , column=1 , padx=5)

btn1.grid(row=0 , column=0 , padx=5 , pady=5)
btn2.grid(row=1 , column=0 , padx=5 , pady=5)
btn3.grid(row=2 , column=0 , padx=5 , pady=5)

btn4.grid(row=0 , column=0 , padx=5 , pady=5)
btn5.grid(row=1 , column=0 , padx=5 , pady=5)
btn6.grid(row=2 , column=0 , padx=5 , pady=5)

root.protocol("WM_DELETE_WINDOW" , exit_app)
root.mainloop()
